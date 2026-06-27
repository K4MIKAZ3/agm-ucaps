"""
Importa UCAPS PROYECTOS Excel a SQL para Supabase.
Uso: python scripts/import_ucaps_excel.py > scripts/import_output.sql
"""
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"d:\Descargas\UCAPS PROYECTOS 2026 JULIO.xlsx")

MUNICIPIO_ALIASES = {
    "BOSCONIA": ("Bosconia", 1),
    "MAGANGUE": ("Magangué", 1),
    "MAGANGUÉ": ("Magangué", 1),
    "ARJONA": ("Arjona", 3),
    "A.CODAZZI": ("Codazzi", 1),
    "CODAZZI": ("Codazzi", 1),
    "AGUACHICA": ("Aguachica", 4),
    "CUCUTA": ("Cúcuta", 2),
    "CÚCUTA": ("Cúcuta", 2),
    "OCANA": ("Ocaña", 4),
    "OCAÑA": ("Ocaña", 4),
    "VILLA DEL ROSARIO": ("Villa del Rosario", 2),
    "SAN CAYETANO": ("San Cayetano", 2),
    "SARDINATA": ("Sardinata", 2),
    "PUERTO SANTANDER": ("Puerto Santander", 2),
    "PTO SANTANDER": ("Puerto Santander", 2),
    "LA GLORIA": ("La Gloria", 4),
    "NEIVA 2025 - 2027": ("Neiva", 3),
    "NEIVA - ORNATO": ("Neiva Ornato", 3),
    "NEIVA ORNATO": ("Neiva Ornato", 3),
    "PITALITO": ("Pitalito", 3),
    "PITALITO GUADUALES": ("Pitalito", 3),
    "LA DORADA": ("La Dorada", 3),
    "EL COPEY": ("El Copey", 1),
    "COPEY": ("El Copey", 1),
    "SAN ESTANISLAO": ("San Estanislao", 1),
    "LOS PATIOS": ("Los Patios", 2),
    "PEREIRA": ("Pereira", 5),
}

SHEET_TO_PROYECTO = {
    "MAGANGUE": "MAGANGUE",
    "MAGANGUE ": "MAGANGUE",
    "BOSCONIA": "BOSCONIA",
    "CUCUTA": "CÚCUTA",
    "SAN CAYETANO": "SAN CAYETANO",
    "PTO SANTANDER": "PUERTO SANTANDER",
    "SARDINATA": "SARDINATA",
    "NEIVA": "NEIVA 2025 - 2027",
    "PITALITO GUADUALES": "PITALITO",
    "LA DORADA": "LA DORADA",
    "AGUACHICA": "AGUACHICA",
    "OCAÑA": "OCAÑA",
    "OCANA": "OCAÑA",
    "LA GLORIA": "LA GLORIA",
    "COPEY": "EL COPEY",
    "SAN ESTANISLAO": "SAN ESTANISLAO",
    "LOS PATIOS": "LOS PATIOS",
    "PEREIRA": "PEREIRA",
    "ORNTO": "NEIVA - ORNATO",
    "ORNATO": "NEIVA - ORNATO",
}

ESTADO_MAP = {
    "EJECUCION": "EJECUCION",
    "EJECUCIÓN": "EJECUCION",
    "PAUSADO": "PAUSADO",
    "NO INICIADO": "NO_INICIADO",
    "FINALIZADO": "FINALIZADO",
    "EN COMPRAS": "EN_COMPRAS",
}


def norm(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def esc(s) -> str:
    if s is None:
        return "NULL"
    text = str(s).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return "'" + text.replace("'", "''") + "'"


def esc_num(n) -> str:
    if n is None or n == "":
        return "0"
    try:
        return str(float(n))
    except (TypeError, ValueError):
        return "0"


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def si_no(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in ("SI", "SÍ", "YES", "TRUE", "1"):
        return True
    if s in ("NO", "FALSE", "0"):
        return False
    return None


def parse_resumen(wb):
    ws = wb["RESUMEN PROYECTOS"]
    rows = []
    for i in range(3, ws.max_row + 1):
        nombre = ws.cell(i, 2).value
        if not nombre or not str(nombre).strip():
            continue
        valor = ws.cell(i, 11).value
        if valor is None or valor == "":
            continue
        rows.append(
            {
                "nombre_corto": str(nombre).strip(),
                "fecha_acta": to_date(ws.cell(i, 3).value),
                "estado_operativo": ws.cell(i, 4).value,
                "ppto_aprobado": si_no(ws.cell(i, 5).value),
                "material_aprobado": si_no(ws.cell(i, 6).value),
                "avance": ws.cell(i, 7).value or 0,
                "estado": ws.cell(i, 8).value,
                "fecha_terminacion": ws.cell(i, 9).value,
                "valor_ucaps": valor,
                "facturado": ws.cell(i, 12).value or 0,
                "ppto_interno": ws.cell(i, 13).value or 0,
            }
        )
    return rows


def find_header_row(ws):
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(14, ws.max_column + 1))]
        joined = " ".join(str(v) for v in row_vals if v).upper()
        if "ACTIVIDADES" in joined and ("ITEMS" in joined or "ÍTEMS" in joined or "TEMS" in joined):
            return r
    return None


def parse_items_sheet(ws):
    header = find_header_row(ws)
    if not header:
        return [], None, None

    municipio = None
    nombre_completo = None
    for r in range(1, header):
        for c in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(r, c).value
            if val is None:
                continue
            s = str(val).strip().upper()
            nxt = ws.cell(r, c + 1).value if c + 1 <= ws.max_column else None
            if s == "MUNICIPIO" and nxt:
                municipio = str(nxt).strip()
            if s == "PROYECTO" and nxt:
                nombre_completo = str(nxt).strip()

    items = []
    for r in range(header + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, min(14, ws.max_column + 1))]
        row_str = " ".join(str(v) for v in row if v).upper()
        if "SUB" in row_str and "TOTAL" in row_str:
            continue
        if "UTILIDAD" in row_str or row_str.startswith("IVA"):
            continue

        item_no = actividad = unidad = None
        cantidad = valor_u = valor_t = instalado = avance = None

        for col in (3, 4):
            v = ws.cell(r, col).value
            act = ws.cell(r, col + 1).value
            uni = ws.cell(r, col + 2).value
            if v is None or act is None:
                continue
            try:
                float(str(v).replace(",", "."))
                if uni and str(act).strip() and not str(act).upper().startswith("SUMINIST"):
                    item_no = v
                    actividad = str(act).strip()
                    unidad = str(uni).strip() if uni else "Und"
                    cantidad = ws.cell(r, col + 3).value
                    valor_u = ws.cell(r, col + 4).value
                    valor_t = ws.cell(r, col + 5).value
                    instalado = ws.cell(r, col + 6).value
                    avance = ws.cell(r, col + 8).value if ws.max_column >= col + 8 else ws.cell(r, col + 7).value
                    break
            except ValueError:
                pass

        if not actividad or str(actividad).upper().startswith("SUMINIST"):
            continue
        if str(instalado).upper() == "ANULADO":
            instalado = 0

        try:
            cantidad_f = float(cantidad or 0)
            valor_u_f = float(valor_u or 0)
            inst_f = float(instalado) if instalado not in (None, "", "ANULADO") else 0
            avance_f = float(avance or 0)
            if avance_f <= 1 and avance_f > 0:
                avance_f *= 100
        except (TypeError, ValueError):
            continue

        items.append(
            {
                "numero_item": int(float(item_no)) if item_no is not None else len(items) + 1,
                "actividad": actividad,
                "unidad": unidad or "Und",
                "cantidad_total": cantidad_f,
                "valor_unitario": valor_u_f,
                "cantidad_ejecutada": inst_f,
                "avance_pct": min(avance_f, 100),
            }
        )
    return items, municipio, nombre_completo


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    resumen = parse_resumen(wb)

    print("-- AGM UCAPS import from Excel")
    print("BEGIN;")

    extra_municipios = {
        v[0]: v[1] for k, v in MUNICIPIO_ALIASES.items() if v[0] not in {
            "Bosconia", "El Copey", "Magangué", "San Estanislao", "Cúcuta",
            "San Cayetano", "Sardinata", "Puerto Santander", "Los Patios",
            "Neiva", "Pitalito", "La Dorada", "Neiva Ornato", "Aguachica",
            "La Gloria", "Ocaña", "Pereira",
        }
    }
    for nombre, zona in extra_municipios.items():
        print(
            f"INSERT INTO municipios (zona_id, nombre) "
            f"SELECT z.id, {esc(nombre)} FROM zonas z WHERE z.codigo = {zona} "
            f"ON CONFLICT (zona_id, nombre) DO NOTHING;"
        )

    for row in resumen:
        key = norm(row["nombre_corto"])
        alias = MUNICIPIO_ALIASES.get(key) or MUNICIPIO_ALIASES.get(key.replace(".", ""))
        if not alias:
            parts = row["nombre_corto"].split()
            alias = (row["nombre_corto"].title(), 3)
        mun_nombre, zona = alias

        estado_key = norm(str(row["estado"] or "NO INICIADO"))
        estado_codigo = ESTADO_MAP.get(row["estado"], ESTADO_MAP.get(estado_key, "NO_INICIADO"))

        avance_pct = float(row["avance"] or 0)
        if avance_pct <= 1:
            avance_pct *= 100

        fecha_acta = esc(to_date(row["fecha_acta"])) if to_date(row["fecha_acta"]) else "NULL"
        fecha_term = row["fecha_terminacion"]
        if isinstance(fecha_term, datetime):
            fecha_term_sql = esc(to_date(fecha_term))
            fecha_nota = "NULL"
        else:
            fecha_term_sql = "NULL"
            fecha_nota = esc(str(fecha_term)) if fecha_term else "NULL"

        ppto = "true" if row["ppto_aprobado"] else "false" if row["ppto_aprobado"] is False else "NULL"
        mat = "true" if row["material_aprobado"] else "false" if row["material_aprobado"] is False else "NULL"

        print(f"""
INSERT INTO proyectos (
  municipio_id, nombre_corto, nombre_completo, fecha_acta, estado_id,
  estado_operativo, ppto_interno_aprobado, material_aprobado,
  valor_ucaps, ppto_interno, facturado, avance_fisico_pct,
  fecha_terminacion, fecha_terminacion_nota, avance_calculado_auto
)
SELECT m.id, {esc(row['nombre_corto'])}, {esc(row['nombre_corto'])}, {fecha_acta}, e.id,
  {esc(str(row['estado_operativo']) if row['estado_operativo'] else None)},
  {ppto}, {mat},
  {esc_num(row['valor_ucaps'])}, {esc_num(row['ppto_interno'])}, {esc_num(row['facturado'])}, {esc_num(avance_pct)},
  {fecha_term_sql}, {fecha_nota}, false
FROM municipios m
JOIN estados_proyecto e ON e.codigo = {esc(estado_codigo)}
WHERE m.nombre = {esc(mun_nombre)}
AND NOT EXISTS (
  SELECT 1 FROM proyectos px WHERE px.nombre_corto = {esc(row['nombre_corto'])}
);
""".strip())

    skip_sheets = {
        "RESUMEN PROYECTOS", "PPTO DE AVANCE", "PAGOS ZONA", "RESUMEN", "Hoja1",
        "Hoja2", "Hoja3", "Hoja4",
    }

    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in skip_sheets or sheet_name[0:1].isdigit() or sheet_name.startswith("2025"):
            continue
        ws = wb[sheet_name]
        items, mun, nombre_completo = parse_items_sheet(ws)
        if not items:
            continue

        proyecto_key = SHEET_TO_PROYECTO.get(sheet_name.strip(), sheet_name.strip())
        print(f"-- Items sheet {sheet_name} -> {proyecto_key} ({len(items)} items)")

        for it in items:
            unidad_codigo = "UND" if norm(it["unidad"]) in ("UND", "UN", "UNIDAD") else "ML" if "ML" in norm(it["unidad"]) else "UND"
            print(f"""
INSERT INTO actividades_catalogo (nombre, categoria_id, unidad_id)
SELECT {esc(it['actividad'])}, ci.id, um.id
FROM categorias_item ci, unidades_medida um
WHERE ci.codigo = 'SUMINISTRO' AND um.codigo = {esc(unidad_codigo)}
AND NOT EXISTS (
  SELECT 1 FROM actividades_catalogo x
  WHERE lower(trim(x.nombre)) = lower(trim({esc(it['actividad'])}))
);
""".strip())

            print(f"""
INSERT INTO proyecto_items (
  proyecto_id, actividad_id, unidad_id, categoria_id, numero_item,
  cantidad_total, valor_unitario, cantidad_ejecutada, avance_pct, orden
)
SELECT p.id, ac.id, um.id, ci.id, {it['numero_item']},
  {esc_num(it['cantidad_total'])}, {esc_num(it['valor_unitario'])},
  {esc_num(it['cantidad_ejecutada'])}, {esc_num(it['avance_pct'])}, {it['numero_item']}
FROM proyectos p
JOIN actividades_catalogo ac ON lower(trim(ac.nombre)) = lower(trim({esc(it['actividad'])}))
JOIN unidades_medida um ON um.codigo = {esc(unidad_codigo)}
JOIN categorias_item ci ON ci.codigo = 'SUMINISTRO'
WHERE p.nombre_corto = {esc(proyecto_key)}
AND NOT EXISTS (
  SELECT 1 FROM proyecto_items pi
  WHERE pi.proyecto_id = p.id AND pi.numero_item = {it['numero_item']}
);
""".strip())

    print("COMMIT;")
    print(f"-- RESUMEN: {len(resumen)} proyectos")


if __name__ == "__main__":
    out_path = Path(__file__).parent / "import_output.sql"
    from contextlib import redirect_stdout

    with out_path.open("w", encoding="utf-8") as f:
        with redirect_stdout(f):
            main()
    print(f"Wrote {out_path} ({out_path.stat().st_size} bytes)", file=sys.stderr)
